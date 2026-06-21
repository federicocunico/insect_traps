#!/usr/bin/env python
"""
Generate a single Excel workbook (experiment_results.xlsx) holding every results
table, plus a selection of confusion-matrix images.

Reuses the parsing / aggregation logic in ``generate_results_tables.py`` so the
numbers stay identical to the .txt / .tex outputs. One sheet per experiment
(EXP1–EXP5) carries the detection (mAP) and classification (P/R/F1) metrics; an
extra "Confusion Matrices" sheet embeds representative normalized matrices.

Run with an env that has pandas + openpyxl + Pillow, e.g.:
    ~/miniconda3/envs/aiprah5090/bin/python generate_results_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

import generate_results_tables as grt

EXPERIMENTS_DIR = grt.EXPERIMENTS_DIR
OUT_PATH = Path("experiment_results.xlsx")

# ── Styling ──────────────────────────────────────────────────────────────────
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
META_FONT = Font(italic=True, size=10, color="444444")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ID_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

ID_COLUMNS = {"Dataset", "Model", "Train", "Test", "Img Size", "Folds"}


def _write_table_sheet(ws, exp_key, df: pd.DataFrame):
    meta = grt.EXPERIMENT_META[exp_key]
    ncols = len(df.columns)
    last_col = get_column_letter(ncols)

    # Title bar
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"{exp_key.upper()}: {meta['title']}"
    c.font, c.fill, c.alignment = TITLE_FONT, TITLE_FILL, CENTER
    ws.row_dimensions[1].height = 24

    # Description + rationale
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = f"Description: {meta['description']}"
    ws["A2"].font, ws["A2"].alignment = META_FONT, WRAP
    ws.row_dimensions[2].height = 42
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].value = f"Rationale: {meta['rationale']}"
    ws["A3"].font, ws["A3"].alignment = META_FONT, WRAP
    ws.row_dimensions[3].height = 42

    # Header row (row 5)
    hdr_row = 5
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=hdr_row, column=j, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = (
            HEADER_FONT, HEADER_FILL, CENTER, BORDER,
        )

    # Data rows, shading on dataset-group change
    prev_ds = None
    shade = False
    for i, (_, row) in enumerate(df.iterrows()):
        r = hdr_row + 1 + i
        cur_ds = row.get("Dataset", None)
        if prev_ds is not None and cur_ds != prev_ds:
            shade = not shade
        prev_ds = cur_ds
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=r, column=j, value=row[col])
            cell.border = BORDER
            if col in ID_COLUMNS:
                cell.alignment = LEFT
                cell.fill = ID_FILL
            else:
                cell.alignment = CENTER
                if shade:
                    cell.fill = PatternFill("solid", fgColor="F2F7FC")

    # Column widths
    for j, col in enumerate(df.columns, start=1):
        body = [len(str(v)) for v in df[col].tolist()]
        width = max([len(str(col))] + body) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(max(width, 9), 22)

    ws.freeze_panes = f"A{hdr_row + 1}"


def _add_confusion_sheet(ws, picks):
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Representative Confusion Matrices (normalized, fold 0)"
    ws["A1"].font, ws["A1"].fill, ws["A1"].alignment = TITLE_FONT, TITLE_FILL, CENTER
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 4

    row = 3
    for label, png in picks:
        if not png.exists():
            continue
        ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=11)
        row += 1
        img = XLImage(str(png))
        # Scale to ~420 px wide, preserving aspect ratio.
        with PILImage.open(png) as p:
            w, h = p.size
        target_w = 420
        scale = target_w / w
        img.width = target_w
        img.height = int(h * scale)
        anchor_cell = f"B{row}"
        ws.add_image(img, anchor_cell)
        # leave enough rows for the image (~18px per row)
        row += int(img.height / 18) + 2


def _confusion_picks(records):
    """Pick one normalized confusion matrix per dataset (exp1 best model) + combos."""
    def cm_path(dir_name):
        return EXPERIMENTS_DIR / dir_name / "fold_0" / "train" / "confusion_matrix_normalized.png"

    candidates = [
        ("EXP1 · Hi-Res · YOLO11m", "exp1_hi_res_yolo11m_fold0"),
        ("EXP1 · Low-Res · YOLO11m", "exp1_low_res_yolo11m_fold0"),
        ("EXP1 · Literature · YOLO11m", "exp1_literature_yolo11m_fold0"),
        ("EXP4 · Hi-Res + Low-Res · YOLO11s", "exp4_hi_res_low_res_yolo11s_fold0"),
        ("EXP4 · All Combined · YOLO11s", "exp4_combined_yolo11s_fold0"),
    ]
    picks = []
    for label, d in candidates:
        p = cm_path(d)
        if p.exists():
            picks.append((label, p))
    return picks


def main():
    print(f"Scanning {EXPERIMENTS_DIR} …")
    records = grt.scan_experiments(EXPERIMENTS_DIR)
    print(f"Found {len(records)} completed results.")
    if not records:
        raise SystemExit("No completed experiments found.")

    builders = {
        "exp1": grt.build_exp1_table,
        "exp2": grt.build_exp2_table,
        "exp3": grt.build_exp3_table,
        "exp4": grt.build_exp4_table,
        "exp5": grt.build_exp5_table,
    }

    wb = Workbook()
    wb.remove(wb.active)

    # Overview sheet
    ov = wb.create_sheet("Overview")
    ov["A1"].value = "EXPERIMENT RESULTS — Insect Trap Detection"
    ov["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ov["A2"].value = f"Generated from: {EXPERIMENTS_DIR}"
    ov["A3"].value = f"Total completed experiments: {len(records)}"
    ov["A3"].font = Font(bold=True)
    for j, h in enumerate(["Sheet", "Experiment", "Rows"], start=1):
        cell = ov.cell(row=5, column=j, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER

    ov_row = 6
    for exp_key, builder in builders.items():
        df = builder(records)
        if df is None or df.empty:
            print(f"  {exp_key}: no data — skipped")
            continue
        ws = wb.create_sheet(exp_key.upper())
        _write_table_sheet(ws, exp_key, df)
        for j, v in enumerate(
            [exp_key.upper(), grt.EXPERIMENT_META[exp_key]["title"], len(df)], start=1
        ):
            cell = ov.cell(row=ov_row, column=j, value=v)
            cell.border = BORDER
            cell.alignment = CENTER if j != 2 else LEFT
        ov_row += 1
        print(f"  {exp_key}: {len(df)} rows -> sheet {exp_key.upper()}")

    ov.column_dimensions["A"].width = 12
    ov.column_dimensions["B"].width = 42
    ov.column_dimensions["C"].width = 8

    # Confusion matrices
    picks = _confusion_picks(records)
    if picks:
        cm_ws = wb.create_sheet("Confusion Matrices")
        _add_confusion_sheet(cm_ws, picks)
        print(f"  confusion matrices: embedded {len(picks)} images")
    else:
        print("  confusion matrices: none found")

    wb.save(OUT_PATH)
    print(f"\nWritten: {OUT_PATH.resolve()}")


if __name__ == "__main__":
    main()
